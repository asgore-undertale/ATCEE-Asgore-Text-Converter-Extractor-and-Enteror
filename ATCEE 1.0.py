#استيراد المكتبات
from PyQt5 import QtCore, QtGui
from PyQt5.QtWidgets import QApplication, QMainWindow, QTextEdit, QCheckBox, QPushButton, QLabel, QMenuBar, QAction, QFileDialog, QMessageBox, QRadioButton
from sys import argv, exit
from os import path, listdir, mkdir
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font

##استيراد السكربتات
from Scripts.Delete_Duplicated_lines import DDL
from Scripts.Sort_lines import Sort
from Scripts.Re_Unshape_Arabic import Reshape
from Scripts.Fit_in_box import fit_in_box, import_from_width_database
from Scripts.Un_Convert import Convert, import_from_converting_database
from Scripts.Reverse_text import Reverse
from Scripts.Extract_from_text import Extract

##
app = QApplication(argv)

##الخطوط
labels_font = QtGui.QFont()
labels_font.setPointSize(9)
textbox_font = QtGui.QFont()
textbox_font.setPointSize(12)

#نافذة عني
AboutWindow = QMainWindow()
AboutWindow.setFixedSize(378, 160)
AboutWindow.setWindowTitle("عني")

about_textbox = QTextEdit(AboutWindow)
about_textbox.setGeometry(QtCore.QRect(0, 0, 378, 160))
about_textbox.setFont(labels_font)
about_textbox.setReadOnly(True)
about_textbox.setText("طوّرت هذه الأداة من قبل Asgore_Undertale\nصفحتي على github:\nhttps://github.com/asgore-undertale\nلك كامل الحرية في التعديل والنشر،\nبشرط ذكري وصفحتي.")


#نافذة خيارات التحويل
OptionsWindow_Width = 400
checkbox_size = [OptionsWindow_Width-5, 16]
textedit_size = [30, 26]
def pos_y(line_num, Height = checkbox_size[1], Between_every_y = 20):
    y = (Between_every_y+checkbox_size[1]) * line_num + (Between_every_y-Height//2)
    return y

OptionsWindow = QMainWindow()
OptionsWindow.setFixedSize(OptionsWindow_Width, 500)
OptionsWindow.setWindowTitle("خيارات التحويل")

DDL_check = QCheckBox("حذف الأسطر المكررة", OptionsWindow)
DDL_check.setGeometry(QtCore.QRect(0, pos_y(0), checkbox_size[0], checkbox_size[1]))
DDL_check.setLayoutDirection(QtCore.Qt.RightToLeft)

SSL_check = QCheckBox("ترتيب السطور من الأقصر للأطول", OptionsWindow)
SSL_check.setGeometry(QtCore.QRect(0, pos_y(1), checkbox_size[0], checkbox_size[1]))
SSL_check.setLayoutDirection(QtCore.Qt.RightToLeft)

SLS_check = QCheckBox("ترتيب السطور من الأطول للأقصر", OptionsWindow)
SLS_check.setGeometry(QtCore.QRect(0, pos_y(2), checkbox_size[0], checkbox_size[1]))
SLS_check.setLayoutDirection(QtCore.Qt.RightToLeft)

RA_check = QCheckBox("تجميد النص العربي", OptionsWindow)
RA_check.setGeometry(QtCore.QRect(0, pos_y(3), checkbox_size[0], checkbox_size[1]))
RA_check.setLayoutDirection(QtCore.Qt.RightToLeft)

UA_check = QCheckBox("إلغاء تجميد النص العربي", OptionsWindow)
UA_check.setGeometry(QtCore.QRect(0, pos_y(4), checkbox_size[0], checkbox_size[1]))
UA_check.setLayoutDirection(QtCore.Qt.RightToLeft)

C_check = QCheckBox("تحويل النص", OptionsWindow)
C_check.setGeometry(QtCore.QRect(0, pos_y(5), checkbox_size[0], checkbox_size[1]))
C_check.setLayoutDirection(QtCore.Qt.RightToLeft)
UC_check = QCheckBox("إلغاء تحويل النص", OptionsWindow)
UC_check.setGeometry(QtCore.QRect(0, pos_y(6), checkbox_size[0], checkbox_size[1]))
UC_check.setLayoutDirection(QtCore.Qt.RightToLeft)

RT_check = QCheckBox("عكس النص", OptionsWindow)
RT_check.setGeometry(QtCore.QRect(0, pos_y(7), checkbox_size[0], checkbox_size[1]))
RT_check.setLayoutDirection(QtCore.Qt.RightToLeft)
RAO_check = QCheckBox("عكس العربية في النص (تجريبي)", OptionsWindow)
RAO_check.setGeometry(QtCore.QRect(0, pos_y(8), checkbox_size[0], checkbox_size[1]))
RAO_check.setLayoutDirection(QtCore.Qt.RightToLeft)

FIB_check = QCheckBox("ضع النص في مربع", OptionsWindow)
FIB_check.setGeometry(QtCore.QRect(0, pos_y(9), checkbox_size[0], checkbox_size[1]))
FIB_check.setLayoutDirection(QtCore.Qt.RightToLeft)

Ext_check = QCheckBox("استخرج من النص", OptionsWindow)
Ext_check.setGeometry(QtCore.QRect(0, pos_y(10), checkbox_size[0], checkbox_size[1]))
Ext_check.setLayoutDirection(QtCore.Qt.RightToLeft)

##
UC_database_button = QPushButton(OptionsWindow)
UC_database_button.setGeometry(QtCore.QRect(35, 10, 93, 56))
UC_database_button.setText("قاعدة بيانات\nالتحويل")
W_database_button = QPushButton(OptionsWindow)
W_database_button.setGeometry(QtCore.QRect(35, 70, 93, 56))
W_database_button.setText("قاعدة بيانات\nعرض الحروف")

start_command = QTextEdit(OptionsWindow)
start_command.setGeometry(QtCore.QRect(10, 142, 50, 26))
start_com_label = QLabel(OptionsWindow)
start_com_label.setGeometry(QtCore.QRect(100, 142, 60, 26))
start_com_label.setText("قبل الأوامر:")
end_command = QTextEdit(OptionsWindow)
end_command.setGeometry(QtCore.QRect(10, 175, 50, 26))
end_com_end_label = QLabel(OptionsWindow)
end_com_end_label.setGeometry(QtCore.QRect(125, 175, 35, 26))
end_com_end_label.setText("بعدها:")

page_command = QTextEdit(OptionsWindow)
page_command.setGeometry(QtCore.QRect(10, 208, 50, 26))
page_com_label = QLabel(OptionsWindow)
page_com_label.setGeometry(QtCore.QRect(65, 208, 95, 26))
page_com_label.setText("أمر صفحة جديدة:")
line_command = QTextEdit(OptionsWindow)
line_command.setGeometry(QtCore.QRect(10, 241, 50, 26))
line_com_label = QLabel(OptionsWindow)
line_com_label.setGeometry(QtCore.QRect(95, 241, 65, 26))
line_com_label.setText("سطر جديد:")

textzone_width = QTextEdit(OptionsWindow)
textzone_width.setGeometry(QtCore.QRect(10, 274, 50, 26))
textzone_width_label = QLabel(OptionsWindow)
textzone_width_label.setGeometry(QtCore.QRect(65, 274, 95, 26))
textzone_width_label.setText("عرض المربع (px):")
textzone_lines = QTextEdit(OptionsWindow)
textzone_lines.setGeometry(QtCore.QRect(10, 307, 50, 26))
textzone_lines_label = QLabel(OptionsWindow)
textzone_lines_label.setGeometry(QtCore.QRect(80, 307, 80, 26))
textzone_lines_label.setText("سطور المربع:")

Slash_check = QCheckBox(u"\u005c\u006e" + ", " + u"\u005c\u0074" + ", " + u"\u005c\u0072" + ", " + u"\u005c\u0061" + "  :مراعاة", OptionsWindow)
Slash_check.setGeometry(QtCore.QRect(15, 335, 140, 26))
Slash_check.setLayoutDirection(QtCore.Qt.RightToLeft)

before_text_convert = QTextEdit(OptionsWindow)
before_text_convert.setGeometry(QtCore.QRect(10, 368, 50, 26))
before_text_convert_label = QLabel(OptionsWindow)
before_text_convert_label.setGeometry(QtCore.QRect(65, 368, 95, 26))
before_text_convert_label.setText("ما قبل النصوص:")
after_text_convert = QTextEdit(OptionsWindow)
after_text_convert.setGeometry(QtCore.QRect(10, 401, 50, 26))
after_text_convert_label = QLabel(OptionsWindow)
after_text_convert_label.setGeometry(QtCore.QRect(65, 401, 95, 26))
after_text_convert_label.setText("ما بعدها:")

min_text_convert = QTextEdit(OptionsWindow)
min_text_convert.setGeometry(QtCore.QRect(10, 434, 50, 26))
min_text_convert_label = QLabel(OptionsWindow)
min_text_convert_label.setGeometry(QtCore.QRect(65, 434, 95, 26))
min_text_convert_label.setText("أقصى حد لقصرها:")
max_text_convert = QTextEdit(OptionsWindow)
max_text_convert.setGeometry(QtCore.QRect(10, 467, 50, 26))
max_text_convert_label = QLabel(OptionsWindow)
max_text_convert_label.setGeometry(QtCore.QRect(65, 467, 95, 26))
max_text_convert_label.setText("أقصى حد لطولها:")
##


#النافذة الرئيسية
MainWindow = QMainWindow()
MainWindow.setFixedSize(756, 344)
MainWindow.setWindowTitle("ATCEE 1.0")

result_text = QTextEdit(MainWindow)
result_text.setGeometry(QtCore.QRect(13, 62, 301, 253))
result_text.setFont(textbox_font)
entered_text = QTextEdit(MainWindow)
entered_text.setGeometry(QtCore.QRect(440, 66, 301, 253))
entered_text.setFont(textbox_font)

convert_button = QPushButton(MainWindow)
convert_button.setGeometry(QtCore.QRect(330, 116, 93, 28))
convert_button.setText("تحويل")
openfile_button = QPushButton(MainWindow)
openfile_button.setGeometry(QtCore.QRect(330, 166, 93, 41))
openfile_button.setText("فتح ملف\nنص")

label = QLabel(MainWindow)
label.setGeometry(QtCore.QRect(654, 36, 81, 20))
label.setFont(labels_font)
label.setText("النص الداخل:")
label_2 = QLabel(MainWindow)
label_2.setGeometry(QtCore.QRect(220, 36, 81, 20))
label_2.setFont(labels_font)
label_2.setText("النص الناتج:")

menubar = QMenuBar(MainWindow)
menubar.setGeometry(QtCore.QRect(0, 0, 756, 26))
menubar.setLayoutDirection(QtCore.Qt.RightToLeft)
converting_options = QAction("خيارات التحويل", MainWindow)
entering = QAction("الإدخال والاستخراج", MainWindow)
about = QAction("عني", MainWindow)
menubar.addAction(converting_options)
menubar.addAction(entering)
menubar.addAction(about)


#نافذة الإدخال
EnteringWindow = QMainWindow()
EnteringWindow.setFixedSize(756, 363)
EnteringWindow.setWindowTitle("نافذة الإدخال")

translate_text = QTextEdit(EnteringWindow)
translate_text.setGeometry(QtCore.QRect(13, 34, 301, 140))
translate_text.setFont(textbox_font)
original_text = QTextEdit(EnteringWindow)
original_text.setGeometry(QtCore.QRect(440, 40, 301, 140))
original_text.setFont(textbox_font)

label = QLabel(EnteringWindow)
label.setGeometry(QtCore.QRect(654, 10, 81, 20))
label.setFont(labels_font)
label.setText("النص الأصلي:")
label_2 = QLabel(EnteringWindow)
label_2.setGeometry(QtCore.QRect(220, 10, 81, 20))
label_2.setFont(labels_font)
label_2.setText("الترجمة:")

enter_button = QPushButton(EnteringWindow)
enter_button.setGeometry(QtCore.QRect(330, 45, 93, 28))
enter_button.setText("إدخال")
convert_enter_button = QPushButton(EnteringWindow)
convert_enter_button.setGeometry(QtCore.QRect(330, 80, 93, 41))
convert_enter_button.setText("تحويل\nوإدخال")
extract_button = QPushButton(EnteringWindow)
extract_button.setGeometry(QtCore.QRect(330, 130, 93, 28))
extract_button.setText("استخراج")

before_text = QTextEdit(EnteringWindow)
before_text.setGeometry(QtCore.QRect(25, 190, 50, 26))
before_text_label = QLabel(EnteringWindow)
before_text_label.setGeometry(QtCore.QRect(70, 190, 60, 26))
before_text_label.setText("ما يسبقه:")
after_text = QTextEdit(EnteringWindow)
after_text.setGeometry(QtCore.QRect(145, 190, 50, 26))
after_text_label = QLabel(EnteringWindow)
after_text_label.setGeometry(QtCore.QRect(215, 190, 160, 26))
after_text_label.setText("ما يلحق النص في الملفات:")

min_text = QTextEdit(EnteringWindow)
min_text.setGeometry(QtCore.QRect(145, 223, 50, 26))
min_text_label = QLabel(EnteringWindow)
min_text_label.setGeometry(QtCore.QRect(175, 223, 200, 26))
min_text_label.setText("أقصى حد لقصر النصوص المستخرجة:")
max_text = QTextEdit(EnteringWindow)
max_text.setGeometry(QtCore.QRect(25, 223, 50, 26))
max_text_label = QLabel(EnteringWindow)
max_text_label.setGeometry(QtCore.QRect(70, 223, 60, 26))
max_text_label.setText("وطولها:")

database_check = QCheckBox("استخدام قاعدة البيانات", EnteringWindow)
database_check.setGeometry(QtCore.QRect(570, 193, 150, 16))
database_check.setLayoutDirection(QtCore.Qt.RightToLeft)
too_long_check = QCheckBox("عدم إدخال ترجمات أطول من النص الأصلي (بقيم الهيكس)", EnteringWindow)
too_long_check.setGeometry(QtCore.QRect(440, 218,280, 16))
too_long_check.setLayoutDirection(QtCore.Qt.RightToLeft)
translation_place_check = QCheckBox(":مكان الترجمة في حال كانت أقصر", EnteringWindow)
translation_place_check.setGeometry(QtCore.QRect(470, 243,250, 16))
translation_place_check.setLayoutDirection(QtCore.Qt.RightToLeft)

first_radio = QRadioButton(EnteringWindow)
first_radio.setGeometry(QtCore.QRect(580, 268,100, 16))
first_radio.setText("أول")
first_radio.setLayoutDirection(QtCore.Qt.RightToLeft)
middle_radio = QRadioButton(EnteringWindow)
middle_radio.setGeometry(QtCore.QRect(510, 268,100, 16))
middle_radio.setText("وسط")
middle_radio.setLayoutDirection(QtCore.Qt.RightToLeft)
last_radio = QRadioButton(EnteringWindow)
last_radio.setGeometry(QtCore.QRect(440, 268,100, 16))
last_radio.setText("آخر")
last_radio.setLayoutDirection(QtCore.Qt.RightToLeft)

input_from_folder = QPushButton(EnteringWindow)
input_from_folder.setGeometry(QtCore.QRect(610, 300, 93, 41))
input_from_folder.setText("المجلد الحاوي\nللملفات")
output_from_folder = QPushButton(EnteringWindow)
output_from_folder.setGeometry(QtCore.QRect(500, 300, 93, 41))
output_from_folder.setText("مجلد الملفات\nبعد الإدخال")
text_database_button = QPushButton(EnteringWindow)
text_database_button.setGeometry(QtCore.QRect(390, 300, 93, 41))
text_database_button.setText("فتح قاعدة\nبيانات النصوص")
extract_database_button = QPushButton(EnteringWindow)
extract_database_button.setGeometry(QtCore.QRect(280, 300, 93, 41))
extract_database_button.setText("فتح قاعدة\nبيانات الاستخراج")


#المتغيرات
converting_database_directory = 'Scripts/Un-Converting_Database.xlsx'
chars_width_database_directory = 'Scripts/Chars_Width_Database.xlsx'
text_database_directory = 'جدول النصوص.xlsx'
extracted_text_database_directory = 'النصوص المستخرجة.xlsx'
input_folder, output_folder = 'المجلد الحاوي للملفات/', 'مجلد الملفات بعد الإدخال/'

if path.exists(converting_database_directory):
    import_from_converting_database(converting_database_directory)
if path.exists(chars_width_database_directory):
    import_from_width_database(chars_width_database_directory)

#الدوال
def open_def(num):
    if num == 0:
        fileName, _ = QFileDialog.getOpenFileName(EnteringWindow, 'قاعدة بيانات النص', '' , '*.xlsx')
        global text_database_directory
        if path.exists(fileName) and fileName != '/' and fileName != text_database_directory:
            text_database_directory = fileName
            QMessageBox.about(EnteringWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 1:
        fileName, _ = QFileDialog.getOpenFileName(OptionsWindow, 'قاعدة بيانات التحويل', '' , '*.xlsx')
        global converting_database_directory
        if path.exists(fileName) and fileName != '/' and fileName != converting_database_directory:
            converting_database_directory = fileName
            import_from_converting_database(converting_database_directory)
            QMessageBox.about(OptionsWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 2:
        fileName, _ = QFileDialog.getOpenFileName(OptionsWindow, 'قاعدة بيانات عرض الحروف', '' , '*.xlsx')
        global chars_width_database_directory
        if path.exists(fileName) and fileName != '/' and fileName != chars_width_database_directory:
            chars_width_database_directory = fileName
            import_from_width_database(chars_width_database_directory)
            QMessageBox.about(OptionsWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 3:
        fileName, _ = QFileDialog.getOpenFileName(EnteringWindow, 'قاعدة بيانات الاستخراج', '' , '*.xlsx')
        global extracted_text_database_directory
        if path.exists(fileName) and fileName != '/' and fileName != extracted_text_database_directory:
            extracted_text_database_directory = fileName
            QMessageBox.about(EnteringWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 4:
        fileName, _ = QFileDialog.getOpenFileName(MainWindow, 'ملف نص', '' , '*')
        if path.exists(fileName) and fileName != '/':
            entered_text.setPlainText(open(fileName, 'r', encoding='utf-8').read())
            QMessageBox.about(MainWindow, "!!تهانيّ", "تم اختيار ملف النص.")
    elif num == 5:
        folder = str(QFileDialog.getExistingDirectory(EnteringWindow, "Select Directory"))+'/'
        global input_folder
        if path.exists(folder) and folder != '/' and folder != input_folder:
            input_folder = folder
            QMessageBox.about(EnteringWindow, "!!تهانيّ", "تم اختيار المجلد.")
    elif num == 6:
        folder = str(QFileDialog.getExistingDirectory(EnteringWindow, "Select Directory"))+'/'
        global output_folder
        if path.exists(folder) and folder != '/' and folder != output_folder:
            output_folder = folder
            QMessageBox.about(EnteringWindow, "!!تهانيّ", "تم اختيار المجلد.")

def convert(text):
    ##إلغاء العملية في حال تحقق إحدى هذه الشروط
    if text == '': return
    if C_check.isChecked() or UC_check.isChecked():
        if not path.exists(converting_database_directory):
            QMessageBox.about(MainWindow, "!!خطأ", "قاعدة بيانات التحويل غير موجودة،\nتم إيقاف كل العمليات.")
            return
    if FIB_check.isChecked():
        if not path.exists(chars_width_database_directory):
            QMessageBox.about(MainWindow, "!!خطأ", "قاعدة بيانات الوضع في مربع غير موجودة،\nتم إيقاف كل العمليات.")
            return
    ##
    
    ##المتغيرات
    if Slash_check.isChecked():
        _start_command = start_command.toPlainText().replace(u'\u005c\u006e', '\n').replace(u'\u005c\u0074', '\t').replace(u'\u005c\u0072', '\r').replace(u'\u005c\u0061', '\a')
        _end_command = end_command.toPlainText().replace(u'\u005c\u006e', '\n').replace(u'\u005c\u0074', '\t').replace(u'\u005c\u0072', '\r').replace(u'\u005c\u0061', '\a')
        _page_command = page_command.toPlainText().replace(u'\u005c\u006e', '\n').replace(u'\u005c\u0074', '\t').replace(u'\u005c\u0072', '\r').replace(u'\u005c\u0061', '\a')
        _line_command = line_command.toPlainText().replace(u'\u005c\u006e', '\n').replace(u'\u005c\u0074', '\t').replace(u'\u005c\u0072', '\r').replace(u'\u005c\u0061', '\a')
    else:
        _start_command = start_command.toPlainText()
        _end_command = end_command.toPlainText()
        _page_command = page_command.toPlainText()
        _line_command = line_command.toPlainText()
    
    if '[b]' in _start_command: _start_command = bytearray.fromhex(_start_command.replace('[b]', '')).decode()
    if '[b]' in _end_command: _end_command = bytearray.fromhex(_end_command.replace('[b]', '')).decode()
    if '[b]' in _page_command: _page_command = bytearray.fromhex(_page_command.replace('[b]', '')).decode()
    if '[b]' in _line_command: _line_command = bytearray.fromhex(_line_command.replace('[b]', '')).decode()
    if '[b]' in textzone_width.toPlainText(): _textzone_width = bytearray.fromhex(textzone_width.toPlainText().replace('[b]', '')).decode()
    else: _textzone_width = textzone_width.toPlainText()
    if '[b]' in textzone_lines.toPlainText(): _textzone_width = bytearray.fromhex(textzone_lines.toPlainText().replace('[b]', '')).decode()
    else: _textzone_lines = textzone_lines.toPlainText()
    if '[b]' in before_text_convert.toPlainText(): _before_text_convert = bytearray.fromhex(before_text_convert.toPlainText().replace('[b]', '')).decode()
    else: _before_text_convert = before_text_convert.toPlainText()
    if '[b]' in after_text_convert.toPlainText(): _after_text_convert = bytearray.fromhex(after_text_convert.toPlainText().replace('[b]', '')).decode()
    else: _after_text_convert = after_text_convert.toPlainText()
    
    ##
    
    if Ext_check.isChecked():#Extract from text
        if _before_text_convert == '' or _after_text_convert == '':
            QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nاملأ حقلي: ما قبل النصوص، ما بعدها.\nعلى الأقل للاستخراج.")
            return
        
        mini = min_text_convert.toPlainText()
        maxi = max_text_convert.toPlainText()
        if '[b]' in mini: mini = bytearray.fromhex(mini.replace('[b]', '')).decode()
        if '[b]' in maxi: maxi = bytearray.fromhex(maxi.replace('[b]', '')).decode()
        if mini == '': mini = 0
        else: mini = int(mini)
        if maxi == '': maxi = 0
        else: maxi = int(maxi)
        
        if mini > maxi:
            QMessageBox.about(EnteringWindow, "!!خطأ", "لا يمكن أن يكون قصر النصوص أطول من طولها.")
            return
        
        text = Extract(text, True, _before_text_convert, _after_text_convert, mini, maxi)
        text = '\n'.join(text)
    
    if DDL_check.isChecked():#Delete Duplicated lines
        text = DDL(text)
    
    if SSL_check.isChecked():#Sort short to long
        text = Sort(text)
    
    if SLS_check.isChecked():#Sort long to short
        text = Sort(text, False)
    
    if RA_check.isChecked() or C_check.isChecked() or FIB_check.isChecked():#Reshape Arabic
        text = Reshape(text)
        
    if FIB_check.isChecked():#Fit in box
        if _textzone_width != '' and _textzone_lines != '':
            text = fit_in_box(text, int(_textzone_width), int(_textzone_lines), _line_command, _page_command, _start_command, _end_command)
        else:
            QMessageBox.about(EnteringWindow, "!!خطأ", "املأ حقلي: عرض المربع، عدد سطور المربع.")

    if C_check.isChecked():#Convert
        text = Convert(text, True, _start_command, _end_command)
    
    if UC_check.isChecked():#Unconvert
        text = Convert(text, False, _start_command, _end_command)
        
    if UA_check.isChecked() or UC_check.isChecked():#Unshape Arabic
        text = Reshape(text, False)
        
    if RT_check.isChecked():#Reverse whole text
        text = Reverse(text, _start_command, _end_command, _page_command, _line_command)
        
    if RAO_check.isChecked():#‫Reverse Arabic only
        text = Reverse(text, _start_command, _end_command, _page_command, _line_command, False)
    
    return text

def enter(convert_bool = True):
    ##المتغيرات
    text_dic = {}
    too_long_dic = {}
    no_found_list = []
    found_list = []
    no_found_log = ''
    too_long_log = ''
    
    ##إلغاء العملية في حال تحقق إحدى هذه الشروط
    if C_check.isChecked() or UC_check.isChecked():
        if not path.exists(converting_database_directory):
            QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nقاعدة بيانات التحويل غير موجودة.")
            return
    if not path.exists(input_folder):
        QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nالمجلد الحاوي للملفات غير موجود.")
        return
    if not path.exists(output_folder):
        mkdir(output_folder) 
        #QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nمجلد الاستخراج غير موجود.")
        #return
    files_list = listdir(input_folder)
    if len(files_list) == 0:
        QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nلا توجد أي ملفات للإدخال إليها.")
        return
    ##
    
    if database_check.isChecked():
        if not path.exists(text_database_directory):
            QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nقاعدة بيانات النصوص غير موجودة.")
            return
        text_xlsx = openpyxl.load_workbook(text_database_directory)
        text_table = text_xlsx.get_sheet_by_name("Main")
        for cell in range(2, len(text_table['A'])+1):
            original_cell_value = text_table['A'+str(cell)].value
            translate_cell_value = text_table['B'+str(cell)].value
            
            if original_cell_value in text_dic:
                if text_dic[original_cell_value] == '' or text_dic[original_cell_value] == None:
                    text_dic[original_cell_value] = translate_cell_value
            else:
                text_dic[original_cell_value] = translate_cell_value
        
        new_d = {}
        for k in sorted(text_dic, key=len, reverse=True):
            new_d[k] = text_dic[k]
        text_dic = new_d
    else:
        if original_text.toPlainText() == '': return
        text_dic[original_text.toPlainText()] = translate_text.toPlainText()
    
    for filename in files_list:
        with open(input_folder+filename, 'rb') as f:
            file_content = f.read()
        
        for text, translation in text_dic.items():
            text = before_text.toPlainText() + text + after_text.toPlainText()
            translation = before_text.toPlainText() + translation + after_text.toPlainText()
            
            if translation_place_check.isChecked() and len(translation.encode('utf-8').hex()) < len(text.encode('utf-8').hex()):
                spaces_count = (len(text.encode('utf-8').hex()) // 2) - (len(translation.encode('utf-8').hex()) // 2)
                if first_radio.isChecked():#first
                    for i in range(spaces_count):
                        translation += ' '
                elif middle_radio.isChecked():#middle
                    for i in range(spaces_count):
                        if i % 2 == 0:
                            translation += ' '
                        else:
                            translation = ' ' + translation
                elif last_radio.isChecked():#last
                    for i in range(spaces_count):
                        translation = ' ' + translation
            
            if bytes(text, 'utf-8') in file_content:
                if convert_bool: translation = convert(translation)
                if too_long_check.isChecked() and len(translation.encode('utf-8').hex()) > len(text.encode('utf-8').hex()):
                    too_long_dic[text] = translation
                else:
                    file_content = file_content.replace(bytes(text, 'utf-8'), bytes(translation, 'utf-8'))
                    found_list.append(text)
                    if text in no_found_list: no_found_list.remove(text)
            else:
                if text not in no_found_log and text not in found_list:
                    no_found_list.append(text)
        
        open(output_folder+filename, 'wb').write(file_content)
    
    for item in no_found_list:
        no_found_log += '> ' + text + '\n'
    for k, v in too_long_dic.items():
        too_long_log += '> ' + k + '\n    ' + k.encode('utf-8').hex() + '\n    ' + v + '\n    ' + v.encode('utf-8').hex() + '\n\n'
    
    if no_found_log == '' and too_long_log == '':
        QMessageBox.about(EnteringWindow, "!!تهانيّ", "انتهى الإدخال.")
    if no_found_log != '':
        msg = QMessageBox()
        msg.setText(no_found_log)
        msg.setWindowTitle("ما لم يتم إيجاده")
        msg.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse);
        msg.exec_()
    if too_long_log != '':
        msg = QMessageBox()
        msg.setText(too_long_log)
        msg.setWindowTitle("أطول من النصوص الأصلية")
        msg.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse);
        msg.exec_()

def extract():
    before = before_text.toPlainText()
    after = after_text.toPlainText()
    if '[b]' in before: before = bytearray.fromhex(before.replace('[b]', '')).decode() 
    if '[b]' in after: after = bytearray.fromhex(after.replace('[b]', '')).decode() 
    if before == '' or after == '':
        QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nاملأ حقلي: ما يسبق النصوص، ما يلحقها.\nعلى الأقل.")
        return
    files_list = listdir(input_folder)
    if len(files_list) == 0:
        QMessageBox.about(EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nلا توجد أي ملفات للاستخراج منها.")
        return
    
    mini = min_text.toPlainText()
    maxi = max_text.toPlainText()
    if '[b]' in mini: mini = bytearray.fromhex(mini.replace('[b]', '')).decode() 
    if '[b]' in maxi: maxi = bytearray.fromhex(maxi.replace('[b]', '')).decode() 
    
    if mini == '': mini = 0
    else: mini = int(mini)
    if maxi == '': maxi = 0
    else: maxi = int(maxi)
    
    if mini > maxi:
        QMessageBox.about(EnteringWindow, "!!خطأ", "لا يمكن أن يكون قصر النصوص أطول من طولها.")
        return
    
    extracted_xlsx = openpyxl.load_workbook(extracted_text_database_directory)
    sheet = extracted_xlsx.get_sheet_by_name("Main")
    row = 2
    '''
    def put_in_sheet(text):
        print(row)
        print(text)
        row += 1
        print(row)
    '''
    sheet.delete_cols(1)
    sheet['A1'].value = "النص الأصلي"
    sheet['A1'].font = Font(bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet['A1'].fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')
    sheet['B2'].value = "لا تفتح هذا الملف أثناء تشغيل الأداة."
    sheet['B2'].font = Font(bold=True)
    
    for filename in files_list:
        with open(input_folder+filename, 'r', encoding="cp437") as f:
            file_content = f.read()
        
        sheet['A'+str(row)].value = filename
        sheet['A'+str(row)].font = Font(bold=True)
        sheet['A'+str(row)].alignment = Alignment(vertical='center', wrap_text=True)
        sheet['A'+str(row)].fill = PatternFill(fill_type='solid', start_color='D112D1', end_color='D112D1')
        row += 1
        
        extracted = Extract(file_content, True, before, after, mini, maxi)
        '''map(put_in_sheet, extracted)'''
        for item in extracted:
            sheet['A'+str(row)].font = Font(bold=False)
            sheet['A'+str(row)].value = item
            row += 1
    
    extracted_xlsx.save(extracted_text_database_directory)
    QMessageBox.about(EnteringWindow, "!!تهانيّ", "انتهى الاستخراج.")

#توصيل الإشارات
convert_button.clicked.connect(lambda: result_text.setPlainText(convert(entered_text.toPlainText())))
openfile_button.clicked.connect(lambda: open_def(4))
converting_options.triggered.connect(lambda: OptionsWindow.show())
entering.triggered.connect(lambda: EnteringWindow.show())
about.triggered.connect(lambda: AboutWindow.show())

text_database_button.clicked.connect(lambda: open_def(0))
UC_database_button.clicked.connect(lambda: open_def(1))
W_database_button.clicked.connect(lambda: open_def(2))
extract_database_button.clicked.connect(lambda: open_def(3))

enter_button.clicked.connect(lambda: enter(False))
extract_button.clicked.connect(lambda: extract())
convert_enter_button.clicked.connect(lambda: enter())
input_from_folder.clicked.connect(lambda: open_def(5))
output_from_folder.clicked.connect(lambda: open_def(6))

#تشغيل البرنامج
MainWindow.show()
exit(app.exec_())