#استيراد المكتبات
from PyQt5.QtWidgets import QApplication, QMessageBox
from sys import argv, exit
from os import path, listdir, mkdir
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font

##استيراد السكربتات
from MainParts.Windows import Windows
from ConvertingScripts.Delete_Duplicated_lines import DDL
from ConvertingScripts.Sort_lines import Sort
from ConvertingScripts.Re_Unshape_Arabic import Reshape
from ConvertingScripts.Fit_in_box import fit_in_box
from ConvertingScripts.Un_Convert import Convert
from ConvertingScripts.Reverse_text import Reverse
from ConvertingScripts.Extract_from_text import Extract
from OtherScripts.Take_From_Table import Take_From_Table
from OtherScripts.Bmfont_to_FIB_Table import BMFont_to_FIB

##
app = QApplication(argv)

#المتغيرات
converting_database_directory = 'ConvertingScripts/Un-Converting_Database.xlsx'
chars_width_database_directory = 'ConvertingScripts/Chars_Width_Database.xlsx'
text_database_directory = 'TextTable.xlsx'
extracted_text_database_directory = 'ExtractedTextTable.xlsx'
input_folder, output_folder = '_FilesFolder/', '_After-enteringFolder/'

convert_database = {}
fit_database = {}

if path.exists(converting_database_directory):
    convert_database = Take_From_Table(converting_database_directory)
if path.exists(chars_width_database_directory):
    fit_database = Take_From_Table(chars_width_database_directory)

Windows()

#الدوال
def BMFont_to_FIB_def():
    BMFont_to_FIB(Windows.BMFont_File.toPlainText() , Windows.FIB_Table.toPlainText())
    QMessageBox.about(Windows.EnteringWindow, "!!تهانيّ", "انتهى العملية.")

def open_def(num):
    if num == 0:
        fileName, _ = QFileDialog.getOpenFileName(Windows.EnteringWindow, 'قاعدة بيانات النص', '' , '*.xlsx')
        global text_database_directory
        if path.exists(fileName) and fileName != '/' and fileName != text_database_directory:
            text_database_directory = fileName
            QMessageBox.about(Windows.EnteringWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 1:
        fileName, _ = QFileDialog.getOpenFileName(Windows.OptionsWindow, 'قاعدة بيانات التحويل', '' , '*.xlsx')
        global converting_database_directory
        if path.exists(fileName) and fileName != '/' and fileName != converting_database_directory:
            converting_database_directory = fileName
            convert_database = Take_From_Table(converting_database_directory)
            QMessageBox.about(Windows.OptionsWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 2:
        fileName, _ = QFileDialog.getOpenFileName(Windows.OptionsWindow, 'قاعدة بيانات عرض الحروف', '' , '*.xlsx')
        global chars_width_database_directory
        if path.exists(fileName) and fileName != '/' and fileName != chars_width_database_directory:
            chars_width_database_directory = fileName
            fit_database = Take_From_Table(chars_width_database_directory)
            QMessageBox.about(Windows.OptionsWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 3:
        fileName, _ = QFileDialog.getOpenFileName(Windows.EnteringWindow, 'قاعدة بيانات الاستخراج', '' , '*.xlsx')
        global extracted_text_database_directory
        if path.exists(fileName) and fileName != '/' and fileName != extracted_text_database_directory:
            extracted_text_database_directory = fileName
            QMessageBox.about(Windows.EnteringWindow, "!!تهانيّ", "تم اختيار قاعدة البيانات.")
    elif num == 4:
        fileName, _ = QFileDialog.getOpenFileName(Windows.MainWindow, 'ملف نص', '' , '*')
        if path.exists(fileName) and fileName != '/':
            entered_text.setPlainText(open(fileName, 'r', encoding='utf-8').read())
            QMessageBox.about(Windows.MainWindow, "!!تهانيّ", "تم اختيار ملف النص.")
    elif num == 5:
        folder = str(QFileDialog.getExistingDirectory(Windows.EnteringWindow, "Select Directory"))+'/'
        global input_folder
        if path.exists(folder) and folder != '/' and folder != input_folder:
            input_folder = folder
            QMessageBox.about(Windows.EnteringWindow, "!!تهانيّ", "تم اختيار المجلد.")
    elif num == 6:
        folder = str(QFileDialog.getExistingDirectory(Windows.EnteringWindow, "Select Directory"))+'/'
        global output_folder
        if path.exists(folder) and folder != '/' and folder != output_folder:
            output_folder = folder
            QMessageBox.about(Windows.EnteringWindow, "!!تهانيّ", "تم اختيار المجلد.")

def convert(text):
    ##إلغاء العملية في حال تحقق إحدى هذه الشروط
    if text == '': return
    if Windows.C_check.isChecked() or Windows.UC_check.isChecked():
        if not path.exists(converting_database_directory):
            QMessageBox.about(Windows.MainWindow, "!!خطأ", "قاعدة بيانات التحويل غير موجودة،\nتم إيقاف كل العمليات.")
            return
    if Windows.FIB_check.isChecked():
        if not path.exists(chars_width_database_directory):
            QMessageBox.about(Windows.MainWindow, "!!خطأ", "قاعدة بيانات عرض الحروف غير موجودة،\nتم إيقاف كل العمليات.")
            return
    ##
    
    ##المتغيرات
    if Windows.Slash_check.isChecked():
        _start_command = Windows.start_command.toPlainText().replace(u'\u005c\u006e', '\n').replace(u'\u005c\u0074', '\t').replace(u'\u005c\u0072', '\r').replace(u'\u005c\u0061', '\a')
        _end_command = Windows.end_command.toPlainText().replace(u'\u005c\u006e', '\n').replace(u'\u005c\u0074', '\t').replace(u'\u005c\u0072', '\r').replace(u'\u005c\u0061', '\a')
        _page_command = Windows.page_command.toPlainText().replace(u'\u005c\u006e', '\n').replace(u'\u005c\u0074', '\t').replace(u'\u005c\u0072', '\r').replace(u'\u005c\u0061', '\a')
        _line_command = Windows.line_command.toPlainText().replace(u'\u005c\u006e', '\n').replace(u'\u005c\u0074', '\t').replace(u'\u005c\u0072', '\r').replace(u'\u005c\u0061', '\a')
    else:
        _start_command = Windows.start_command.toPlainText()
        _end_command = Windows.end_command.toPlainText()
        _page_command = Windows.page_command.toPlainText()
        _line_command = Windows.line_command.toPlainText()
    
    if '[b]' in _start_command: _start_command = bytearray.fromhex(_start_command.replace('[b]', '')).decode()
    if '[b]' in _end_command: _end_command = bytearray.fromhex(_end_command.replace('[b]', '')).decode()
    if '[b]' in _page_command: _page_command = bytearray.fromhex(_page_command.replace('[b]', '')).decode()
    if '[b]' in _line_command: _line_command = bytearray.fromhex(_line_command.replace('[b]', '')).decode()
    if '[b]' in Windows.textzone_width.toPlainText(): _textzone_width = bytearray.fromhex(Windows.textzone_width.toPlainText().replace('[b]', '')).decode()
    else: _textzone_width = Windows.textzone_width.toPlainText()
    if '[b]' in Windows.textzone_lines.toPlainText(): _textzone_width = bytearray.fromhex(Windows.textzone_lines.toPlainText().replace('[b]', '')).decode()
    else: _textzone_lines = Windows.textzone_lines.toPlainText()
    if '[b]' in Windows.before_text_convert.toPlainText(): _before_text_convert = bytearray.fromhex(Windows.before_text_convert.toPlainText().replace('[b]', '')).decode()
    else: _before_text_convert = Windows.before_text_convert.toPlainText()
    if '[b]' in Windows.after_text_convert.toPlainText(): _after_text_convert = bytearray.fromhex(Windows.after_text_convert.toPlainText().replace('[b]', '')).decode()
    else: _after_text_convert = Windows.after_text_convert.toPlainText()
    
    ##
    
    if Windows.Ext_check.isChecked():#Extract from text
        if _before_text_convert == '' or _after_text_convert == '':
            QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nاملأ حقلي: ما قبل النصوص، ما بعدها.\nعلى الأقل للاستخراج.")
            return
        
        mini = Windows.min_text_convert.toPlainText()
        maxi = Windows.max_text_convert.toPlainText()
        if '[b]' in mini: mini = bytearray.fromhex(mini.replace('[b]', '')).decode()
        if '[b]' in maxi: maxi = bytearray.fromhex(maxi.replace('[b]', '')).decode()
        if mini == '': mini = 0
        else: mini = int(mini)
        if maxi == '': maxi = 0
        else: maxi = int(maxi)
        
        if mini > maxi:
            QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "لا يمكن أن يكون قصر النصوص أطول من طولها.")
            return
        
        text = Extract(text, True, _before_text_convert, _after_text_convert, mini, maxi)
        text = '\n'.join(text)
    
    if Windows.DDL_check.isChecked():#Delete Duplicated lines
        text = DDL(text)
    
    if Windows.SSL_check.isChecked():#Sort short to long
        text = Sort(text)
    
    if Windows.SLS_check.isChecked():#Sort long to short
        text = Sort(text, False)
    
    if Windows.RA_check.isChecked() or Windows.C_check.isChecked() or Windows.FIB_check.isChecked():#Reshape Arabic
        text = Reshape(text)
        
    if Windows.FIB_check.isChecked():#Fit in box
        if _textzone_width != '' and _textzone_lines != '':
            text = fit_in_box(text, fit_database, int(_textzone_width), int(_textzone_lines), _line_command, _page_command, _start_command, _end_command)
        else:
            QMessageBox.about(EnteringWindow, "!!خطأ", "املأ حقلي: عرض المربع، عدد سطور المربع.")

    if Windows.C_check.isChecked():#Convert
        text = Convert(text, convert_database, True, _start_command, _end_command)
    
    if Windows.UC_check.isChecked():#Unconvert
        text = Convert(text, convert_database, False, _start_command, _end_command)
        
    if Windows.UA_check.isChecked() or Windows.UC_check.isChecked():#Unshape Arabic
        text = Reshape(text, False)
        
    if Windows.RT_check.isChecked():#Reverse whole text
        text = Reverse(text, _start_command, _end_command, _page_command, _line_command)
        
    if Windows.RAO_check.isChecked():#‫Reverse Arabic only
        text = Reverse(text, _start_command, _end_command, _page_command, _line_command, False)
    return text

def enter(convert_bool = True):
    ##المتغيرات
    text_dic = {}
    too_long_dic = {}
    no_found_list = []
    found_list = []
    no_found_log = ''
    too_long_log = ''
    
    ##إلغاء العملية في حال تحقق إحدى هذه الشروط
    if Windows.C_check.isChecked() or Windows.UC_check.isChecked():
        if not path.exists(converting_database_directory):
            QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nقاعدة بيانات التحويل غير موجودة.")
            return
    if not path.exists(input_folder):
        QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nالمجلد الحاوي للملفات غير موجود.")
        return
    if not path.exists(output_folder):
        mkdir(output_folder)
    files_list = listdir(input_folder)
    if len(files_list) == 0:
        QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nلا توجد أي ملفات للإدخال إليها.")
        return
    ##
    
    if Windows.database_check.isChecked():
        if not path.exists(text_database_directory):
            QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "تم إيقاف كل العمليات،\nقاعدة بيانات النصوص غير موجودة.")
            return
        text_xlsx = openpyxl.load_workbook(text_database_directory)
        text_table = text_xlsx.get_sheet_by_name("Main")
        for cell in range(2, len(text_table['A'])+1):
            original_cell_value = text_table['A'+str(cell)].value
            translate_cell_value = text_table['B'+str(cell)].value
            
            if original_cell_value in text_dic:
                if text_dic[original_cell_value] == '' or text_dic[original_cell_value] == None:
                    text_dic[original_cell_value] = translate_cell_value
            else:
                text_dic[original_cell_value] = translate_cell_value
        
        new_d = {}
        for k in sorted(text_dic, key=len, reverse=True):
            new_d[k] = text_dic[k]
        text_dic = new_d
    else:
        if Windows.original_text.toPlainText() == '': return
        text_dic[Windows.original_text.toPlainText()] = Windows.translate_text.toPlainText()
    
    for filename in files_list:
        with open(input_folder+filename, 'rb') as f:
            file_content = f.read()
        
        for text, translation in text_dic.items():
            text = Windows.before_text.toPlainText() + text + Windows.after_text.toPlainText()
            translation = Windows.before_text.toPlainText() + translation + Windows.after_text.toPlainText()
            
            if Windows.translation_place_check.isChecked() and len(translation.encode('utf-8').hex()) < len(text.encode('utf-8').hex()):
                spaces_count = (len(text.encode('utf-8').hex()) // 2) - (len(translation.encode('utf-8').hex()) // 2)
                if Windows.first_radio.isChecked():#first
                    for i in range(spaces_count):
                        translation += ' '
                elif Windows.middle_radio.isChecked():#middle
                    for i in range(spaces_count):
                        if i % 2 == 0:
                            translation += ' '
                        else:
                            translation = ' ' + translation
                elif Windows.last_radio.isChecked():#last
                    for i in range(spaces_count):
                        translation = ' ' + translation
            
            if bytes(text, 'utf-8') in file_content:
                if convert_bool: translation = convert(translation)
                if Windows.too_long_check.isChecked() and len(translation.encode('utf-8').hex()) > len(text.encode('utf-8').hex()):
                    too_long_dic[text] = translation
                else:
                    file_content = file_content.replace(bytes(text, 'utf-8'), bytes(translation, 'utf-8'))
                    found_list.append(text)
                    if text in no_found_list: no_found_list.remove(text)
            else:
                if text not in no_found_log and text not in found_list:
                    no_found_list.append(text)
        
        open(output_folder+filename, 'wb').write(file_content)
    
    for item in no_found_list:
        no_found_log += '> ' + text + '\n'
    for k, v in too_long_dic.items():
        too_long_log += '> ' + k + '\n    ' + k.encode('utf-8').hex() + '\n    ' + v + '\n    ' + v.encode('utf-8').hex() + '\n\n'
    
    if no_found_log == '' and too_long_log == '':
        QMessageBox.about(Windows.EnteringWindow, "!!تهانيّ", "انتهى الإدخال.")
    if no_found_log != '':
        msg = QMessageBox()
        msg.setText(no_found_log)
        msg.setWindowTitle("ما لم يتم إيجاده")
        msg.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse);
        msg.exec_()
    if too_long_log != '':
        msg = QMessageBox()
        msg.setText(too_long_log)
        msg.setWindowTitle("أطول من النصوص الأصلية")
        msg.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse);
        msg.exec_()

def extract():
    before = Windows.before_text.toPlainText()
    after = Windows.after_text.toPlainText()
    if '[b]' in before: before = bytearray.fromhex(before.replace('[b]', '')).decode() 
    if '[b]' in after: after = bytearray.fromhex(after.replace('[b]', '')).decode() 
    if before == '' or after == '':
        QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nاملأ حقلي: ما يسبق النصوص، ما يلحقها.\nعلى الأقل.")
        return
    files_list = listdir(input_folder)
    if len(files_list) == 0:
        QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "تم إيقاف العملية،\nلا توجد أي ملفات للاستخراج منها.")
        return
    
    mini = Windows.min_text.toPlainText()
    maxi = Windows.max_text.toPlainText()
    if '[b]' in mini: mini = bytearray.fromhex(mini.replace('[b]', '')).decode() 
    if '[b]' in maxi: maxi = bytearray.fromhex(maxi.replace('[b]', '')).decode() 
    
    if mini == '': mini = 0
    else: mini = int(mini)
    if maxi == '': maxi = 0
    else: maxi = int(maxi)
    
    if mini > maxi:
        QMessageBox.about(Windows.EnteringWindow, "!!خطأ", "لا يمكن أن يكون قصر النصوص أطول من طولها.")
        return
    
    extracted_xlsx = openpyxl.load_workbook(extracted_text_database_directory)
    sheet = extracted_xlsx.get_sheet_by_name("Main")
    row = 2
    '''
    def put_in_sheet(text):
        print(row)
        print(text)
        row += 1
        print(row)
    '''
    sheet.delete_cols(1)
    sheet['A1'].value = "النص الأصلي"
    sheet['A1'].font = Font(bold=True)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet['A1'].fill = PatternFill(fill_type='solid', start_color='ff8327', end_color='ff8327')
    sheet['B2'].value = "لا تفتح هذا الملف أثناء تشغيل الأداة."
    sheet['B2'].font = Font(bold=True)
    
    for filename in files_list:
        with open(input_folder+filename, 'r', encoding="cp437") as f:
            file_content = f.read()
        
        sheet['A'+str(row)].value = filename
        sheet['A'+str(row)].font = Font(bold=True)
        sheet['A'+str(row)].alignment = Alignment(vertical='center', wrap_text=True)
        sheet['A'+str(row)].fill = PatternFill(fill_type='solid', start_color='D112D1', end_color='D112D1')
        row += 1
        
        extracted = Extract(file_content, True, before, after, mini, maxi)
        '''map(put_in_sheet, extracted)'''
        for item in extracted:
            sheet['A'+str(row)].font = Font(bold=False)
            sheet['A'+str(row)].value = item
            row += 1
    
    extracted_xlsx.save(extracted_text_database_directory)
    QMessageBox.about(Windows.EnteringWindow, "!!تهانيّ", "انتهى الاستخراج.")

#توصيل الإشارات
Windows.convert_button.clicked.connect(lambda: Windows.result_text.setPlainText(convert(Windows.entered_text.toPlainText())))
Windows.convert_button.clicked.connect(lambda: Windows.result_text.setPlainText(convert(Windows.entered_text.toPlainText())))
Windows.openfile_button.clicked.connect(lambda: open_def(4))
Windows.converting_options.triggered.connect(lambda: Windows.OptionsWindow.show())
Windows.entering.triggered.connect(lambda: Windows.EnteringWindow.show())
Windows.about.triggered.connect(lambda: Windows.AboutWindow.show())
Windows.important.triggered.connect(lambda: Windows.ImportantWindow.show())

Windows.text_database_button.clicked.connect(lambda: open_def(0))
Windows.UC_database_button.clicked.connect(lambda: open_def(1))
Windows.W_database_button.clicked.connect(lambda: open_def(2))
Windows.extract_database_button.clicked.connect(lambda: open_def(3))

Windows.enter_button.clicked.connect(lambda: enter(False))
Windows.extract_button.clicked.connect(lambda: extract())
Windows.convert_enter_button.clicked.connect(lambda: enter())
Windows.input_from_folder.clicked.connect(lambda: open_def(5))
Windows.output_from_folder.clicked.connect(lambda: open_def(6))

Windows.BMFont_to_FIB_Action.triggered.connect(lambda: Windows.BMFont_to_FIB_Window.show())
Windows.BMFont_Button.clicked.connect(lambda: BMFont_to_FIB_def())

#تشغيل البرنامج
Windows.MainWindow.show()
exit(app.exec_())